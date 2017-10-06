#!/usr/bin/env python2.7
""" Checks CLM XML against given Olympic limits """

import logging
import operator
import os
import re
import shlex
import sys

if (sys.version_info[0] != 2) or (sys.version_info[1] < 7):
    print >> sys.stderr, ("%s: Error: This script requires Python version " +
                          ">= 2.7 and < 3.0 (%d.%d found)") % \
        (os.path.basename(__file__), sys.version_info[0], sys.version_info[1])
    sys.exit(1)
else:
    from argparse import ArgumentParser, RawDescriptionHelpFormatter
    from collections import OrderedDict

try:
    import openpyxl
except:
    print >> sys.stderr, "This script requires \"openpyxl\" Python module"
    sys.exit(1)
try:
    import xlrd
except:
    print >> sys.stderr, "This script requires \"xlrd\" Python module"
    sys.exit(1)
sys.path.append(os.path.join(os.path.dirname(os.path.realpath(__file__)), "..",
                             "ClmCompiler"))
try:
    import ClmCompiler
except:
    print >> sys.stderr, ("%s: Error: ClmCompiler.py not found. Please " +
                          "place it to the same directory as this script") % \
        os.path.basename(__file__)
    sys.exit(1)


class MyArgumentParser(ArgumentParser):
    """ ArgumentParser that can read switch and argument from same parameter
    file line
    """
    def convert_arg_line_to_args(self, arg_line):
        """ Converts content of parameter file line to list of command line
        arguments

        Arguments:Line from argument file
        Returns vector of argument sin line
        """
        lexer = shlex.shlex(
            arg_line.replace(chr(0x93), '"').replace(chr(0x94), '"').
            replace(chr(0x91), "'").replace(chr(0x92), "'"),
            posix=True)
        lexer.wordchars += "-,;"
        return list(lexer)


def fatal(errmsg):
    """ Prints given fatal error message and exits """
    logging.fatal(errmsg)
    sys.exit(1)


def fatal_if(cond, errmsg):
    """ If condition is true - prints given error message and exits """
    if not cond:
        return
    logging.fatal(errmsg)
    sys.exit(1)


class LoggingRegisteringFilter(logging.Filter):
    """ Logging filter that registers how many numbers and warnings were
    printed
    """
    num_errors = 0
    num_warnings = 0

    def filter(self, record):
        """ Updates number of records/warnings according to given record """
        if record.levelno == logging.ERROR:
            LoggingRegisteringFilter.num_errors += 1
        elif record.levelno == logging.WARNING:
            LoggingRegisteringFilter.num_warnings += 1
        return True


class WorksheetBase(object):
    """ Base interface/utility class for worksheet objects

    Public read-only attributes
    nrows -- Number of rows
    ncols -- Number of columns
    name  -- Worksheet name

    Protected attributes
    _nrows -- Number of rows
    _ncols -- Number of columns
    _name  -- Worksheet name
    """
    def __init__(self, nrows=0, ncols=0, name=""):
        """ Constructor

        Arguments:
        nrows -- Number of rows
        ncols -- Number of columns
        name  -- Worksheet name
        """
        self._nrows = nrows
        self._ncols = ncols
        self._name = name

    def __getattr__(self, name):
        """ Returns read-only attribute of given  name """
        if hasattr(self, "_" + name):
            return getattr(self, "_" + name)
        raise AttributeError("Unknown attribute %s" % name)

    def get_cell(self, row, column):
        """ Pure virtual function - returns cell text value

        Arguments:
        row    -- 1-based row index
        column -- 1-based column index or Excel-style column name
        Returns cell text
        """
        raise NotImplementedError(
            "get_value() shall be implemented in derived class")

    def toascii(self, s):
        """ Returns ASCII representation of given object """
        if s is None:
            return ""
        if isinstance(s, unicode):
            return s.encode('utf-8')
        return str(s)

    @staticmethod
    def col_alpha_to_num(col_name):
        """ Converts column name to 1-based column index """
        ret = 0
        for c in col_name.upper():
            ret = (ret * 26) + (ord(c) - 'A') + 1
        return ret

    @staticmethod
    def col_num_to_alpha(col_num):
        """ Converts 1-based column index to column name """
        ret = ""
        while col_num > 0:
            col_num -= 1
            ret = chr(ord('A') + (col_num % 26)) + ret
            col_num /= 26
        return ret

    @staticmethod
    def cell_name(row, column):
        """ Returns excel-style cell name (address) """
        return "%s%d" % \
            (WorksheetBase.col_num_to_alpha(column) if
             isinstance(column, int) else column, row)


class WorksheetXlrd(WorksheetBase):
    """ Xlrd-based (.xls) worksheet implementation """
    def __init__(self, worksheet):
        """ Constructor

        Arguments:
        worksheet -- xlrd worksheet object
        """
        WorksheetBase.__init__(
            self, nrows=worksheet.nrows, ncols=worksheet.ncols,
            name=worksheet.name)
        self._worksheet = worksheet

    def get_cell(self, row, column):
        """ Returns cell text value

        Arguments:
        row    -- 1-based row index
        column -- 1-based column index or Excel-style column name
        Returns cell text
        """
        if not isinstance(column, int):
            column = WorksheetBase.col_alpha_to_num(column)
        return self.toascii(self._worksheet.cell(row - 1, column - 1).value) \
            .strip()


class WorksheetOpenpyxl(WorksheetBase):
    """ Openpyxl-based (.xlsx) worksheet implementation """
    def __init__(self, worksheet):
        """ Constructor

        Arguments:
        worksheet -- openpyxl worksheet object
        """
        WorksheetBase.__init__(
            self, nrows=worksheet.max_row, ncols=worksheet.max_column,
            name=worksheet.title)
        self._worksheet = worksheet

    def get_cell(self, row, column):
        """ Returns cell text value

        Arguments:
        row    -- 1-based row index
        column -- 1-based column index or Excel-style column name
        Returns cell text
        """
        if not isinstance(column, int):
            column = WorksheetBase.col_alpha_to_num(column)
        cell = self._worksheet.cell(row=row, column=column).value
        return self.toascii(cell).strip()


class WorkbookBase(object):
    """ Base interface/utility class for workbook objects

    Public read-only attributes
    filename    -- Workbook file name
    sheet_names -- List of sheet names
    nsheets     -- Number of sheets

    Protected attributes
    _filename    -- Workbook file name
    _sheet_names -- List of sheet names
    """
    def __init__(self, filename, sheet_names):
        """ Constructor

        Arguments:
        filename    -- Workbook file name
        sheet_names -- List of sheet names
        """
        self._filename = filename
        self._sheet_names = sheet_names

    def __getattr__(self, name):
        """ Returns read-only attribute of given  name """
        if hasattr(self, "_" + name):
            return getattr(self, "_" + name)
        if name == "nsheets":
            return len(self._sheet_names)
        raise AttributeError("Unknown attribute %s" % name)

    def sheet_by_name(self, name):
        """ Pure virtual function - returns worksheet by name

        Argument:
        name -- Worksheet name
        Returns object derived from WorksheetBase or None
        """
        raise NotImplementedError(
            "sheet_by_name() shall be implemented in derived class")

    def sheet_by_index(self, index):
        """ Pure virtual function - returns worksheet by index

        Argument:
        index -- 0-based worksheet index
        Returns object derived from WorksheetBase or None
        """
        raise NotImplementedError(
            "sheet_by_index() shall be implemented in derived class")


class WorkbookXlrd(WorkbookBase):
    """ Xlrd-based (.xls) workbook implementation """
    def __init__(self, filename):
        """ Constructor

        Arguments:
        filename -- Workbook file name
        """
        self._workbook = xlrd.open_workbook(filename=filename)
        WorkbookBase.__init__(self, filename=filename,
                              sheet_names=self._workbook.sheet_names())

    def sheet_by_name(self, name):
        """ Returns worksheet by name

        Argument:
        name -- Worksheet name
        Returns object derived from WorksheetBase or None
        """
        return WorksheetXlrd(self._workbook.sheet_by_name(name)) \
            if name in self.sheet_names else None

    def sheet_by_index(self, index):
        """ Returns worksheet by index

        Argument:
        index -- 0-based worksheet index
        Returns object derived from WorksheetBase or None
        """
        return WorksheetXlrd(self._workbook.sheet_by_index(index)) \
            if 0 <= index < self.nsheets else None


class WorkbookOpenpyxl(WorkbookBase):
    """ Openpyxl-based (.xlsx) workbook implementation """
    def __init__(self, filename):
        """ Constructor

        Arguments:
        filename -- Workbook file name
        """
        self._workbook = openpyxl.load_workbook(filename=filename,
                                                read_only=True,
                                                data_only=True)
        WorkbookBase.__init__(self, filename=filename,
                              sheet_names=self._workbook.sheetnames)

    def sheet_by_name(self, name):
        """ Returns worksheet by name

        Argument:
        name -- Worksheet name
        Returns object derived from WorksheetBase or None
        """
        return WorksheetOpenpyxl(self._workbook.get_sheet_by_name(name)) \
            if name in self.sheet_names else None

    def sheet_by_index(self, index):
        """ Returns worksheet by index

        Argument:
        index -- 0-based worksheet index
        Returns object derived from WorksheetBase or None
        """
        return WorksheetOpenpyxl(self._workbook.worksheets[index]) \
            if 0 <= index < self.nsheets else None


def open_workbook(filename):
    """ Returns Workbook object for given workbook file """
    fatal_if(not os.path.isfile(filename),
             "Can't find spreadsheet file \"%s\"" % filename)
    if os.path.splitext(filename)[1].lower() in (".xls", ".xlsm"):
        return WorkbookXlrd(filename)
    if os.path.splitext(filename)[1].lower() == ".xlsx":
        return WorkbookOpenpyxl(filename)
    fatal("Only .xls, .xlsm and .xlsx spreadsheet files are supported")


class ColumnSignature(object):
    """ Column signature (essential parameters from power column heading and
    rule table row)

    Attributes:
    chains        -- Chain count
    bw            -- Bandwidth (member of ClmCompiler.Bandwidth)
    rate_type     -- Rate type (member of ClmCompiler.RateInfo.RateType)
    is_cdd        -- True for CDD rates
    is_stbc       -- True for STBC rates
    is_sdm        -- True for SDM rates
    txbf          -- True for TXBF rates, false for non-TXBF rates, None if not
                     specified
    default_rates -- Frozenset of rates that correspond to given column
                     signature at face value
    """
    def __init__(self, heading, error_prefix):
        """ Constructor

        Arguments:
        heading      -- Power column or rule row heading
        error_prefix -- Prefix to use in error messages. Shall identify where
                        from heading was obtained
        """
        if "SISO" in heading:
            self.chains = 1
        else:
            m = re.search(r"\b(\d)\s*Tx\b", heading,
                          re.IGNORECASE)
            fatal_if(not m, "%sCan't determine chain count" % error_prefix)
            self.chains = int(m.group(1))
        m = re.search(
            r"\b((?P<ht>HT(?P<htbw>20|40|80|160))|" +
            r"(?P<vht>VHT(?P<vhtbw>20|40|80|160))|" +
            r"(?P<ofdm>(a|g|OFDM))|(?P<dsss>b|DSSS))\b",
            heading)
        fatal_if(not m, "%sRate type not found" % error_prefix)
        if m.group("ht"):
            self.bw = ClmCompiler.Bandwidth.parse(m.group("htbw"))
            if "DSSS" in heading:
                self.rate_type = ClmCompiler.RateInfo.RateType.DSSS
            elif "OFDM" in heading:
                self.rate_type = ClmCompiler.RateInfo.RateType.OFDM
            else:
                self.rate_type = ClmCompiler.RateInfo.RateType.MCS
        elif m.group("vht"):
            self.bw = ClmCompiler.Bandwidth.parse(m.group("vhtbw"))
            self.rate_type = ClmCompiler.RateInfo.RateType.VHT
        elif m.group("ofdm"):
            self.bw = ClmCompiler.Bandwidth._20
            self.rate_type = ClmCompiler.RateInfo.RateType.OFDM
        elif m.group("dsss"):
            self.bw = ClmCompiler.Bandwidth._20
            self.rate_type = ClmCompiler.RateInfo.RateType.DSSS
        self.is_cdd = re.search(r"\bCDD\b", heading) is not None
        self.is_stbc = re.search(r"\bSTBC\b", heading) is not None
        self.is_sdm = re.search(r"\bSDM\b", heading) is not None
        m = re.search(r"\b(non(-|\s)?)?TXBF\b", heading,
                      re.IGNORECASE)
        self.txbf = (not bool(m.group(1))) if m else None
        self.default_rates = self._get_default_rates()

    def __eq__(self, other):
        """ Equality comparison """
        return (type(self) is type(other)) and \
               (self.__dict__ == other.__dict__)

    def __ne__(self, other):
        """ Inequality comparison """
        return not self.__eq__(other)

    def __hash__(self):
        """ Hash for using object as dictionary key """
        return sum(v.__hash__() for v in self.__dict__.values())

    def _get_default_rates(self):
        """ Returns default rates set for column """
        ret = set()
        for rate in ClmCompiler.RatesInfo.get_rates_by_chains(self.chains):
            if (not rate.singular) or (rate.rate_type != self.rate_type):
                continue
            skip = True
            for mod_value, mod_repr in [(self.is_cdd, "CDD"),
                                        (self.is_stbc, "STBC"),
                                        (self.txbf, "TXBF")]:
                if (mod_value is not None) and \
                        (bool(mod_value) != (mod_repr in rate.name)):
                    break
                if self.is_sdm and \
                        ((self.chains == 1) or
                         (self.rate_type not in
                          (ClmCompiler.RateInfo.RateType.MCS,
                           ClmCompiler.RateInfo.RateType.VHT)) or
                         any([exp in rate.name
                              for exp in ("CDD", "STBC", "TXBF")])):
                    break
            else:
                skip = False
            if skip:
                continue
            ret.add(rate)
        return frozenset(ret)

    def __repr__(self):
        """ Debugger representation """
        return "<%s>" % \
            ", ".join(rg for rg in
                      sorted(set([rate.group for rate in self.default_rates])))


class ColumnRules(object):
    """ Interpreter of "Rule" page and container of its data

    Private attributes:
    _rules -- Container of rules, mapping column headings (page-specific and
              clarified by channels) to rate sets that shall be used for those
              columns instead of default rate sets. Multilevel dictionary,
              indexed by page names then by band then by channel ranges (for
              whole-band rules this level contains the only index: None), then
              by column signature. Leaves are RuleInfo objects
    """
    # Name of page that contains rules in template spreadsheet
    RULE_SHEET_NAMES = ("Rule", "Rules", "Global Rules")

    class RuleInfo(object):
        """ Leaf information about one rule row

        Attributes:
        signature_text -- Text representation of signature, with newlines
                          replaced by spaces
        row            -- Spreadsheet row number where rule defined
        rates          -- Set of redress rates (ClmCompiler.RateInfo objects)
                          defined in rule
        """
        def __init__(self, signature_text, row):
            """ Constructor

            Arguments:
            signature_text -- Text representation of signature, with newlines
                              replaced by spaces
            row            -- Spreadsheet row number where rule defined
            """
            self.signature_text = re.sub(r"\s+", " ", signature_text).strip()
            self.row = row
            self.rates = set()

        def add_rate(self, rate_info, duplicating_rate_groups=None):
            """ Adds rate to set of redress rates

            Arguments:
            rate_info               -- Rate (ClmCompiler.RateInfo object) to
                                       add
            duplicating_rate_groups -- Optional output parameter - set of rate
                                       groups of rates, added more than once
            """
            if rate_info in self.rates:
                if duplicating_rate_groups:
                    duplicating_rate_groups.add(rate_info.group)
            else:
                self.rates.add(rate_info)

    def __init__(self, ws=None):
        """ Constructor

        Arguments:
        ws -- Worksheet object if found in template spreadsheet, otherwise None
        """
        self._rules = {}
        if ws is None:
            return
        ST_COMMENT, ST_HEADING, ST_COLNAMES, ST_RULE = range(4)
        state = ST_COMMENT
        current_indent = 0
        rule_collection = {}
        for row in range(1, ws.nrows + 1):
            row_error_prefix = "Sheet \"%s\", row %d: " % (ws.name, row)
            for new_indent in range(ws.ncols):
                if ws.get_cell(row, 1 + new_indent):
                    break
            else:
                new_indent = current_indent
            if state == ST_COMMENT:
                if ws.get_cell(row, 1 + new_indent) == "Pages":
                    state = ST_HEADING
                    current_indent = new_indent
            elif state == ST_HEADING:
                if ws.get_cell(row, 1 + current_indent) == "Channels":
                    state = ST_COLNAMES
                else:
                    fatal(("%sHeading row not followed by " +
                           "Channels/Column/Rates row") % row_error_prefix)
            elif state in (ST_COLNAMES, ST_RULE):
                cell = ws.get_cell(row, 1 + current_indent)
                if cell:
                    state = ST_RULE
                    if cell.startswith(":"):
                        continue
                else:
                    state = ST_COMMENT
            if state == ST_HEADING:
                rule_collection = {}
                for country in \
                    [c.strip() for c in
                     re.sub(r"\s", " ",
                            ws.get_cell(row, 2 + new_indent)).split(",")]:
                    if country == "":
                        continue
                    if country in self._rules:
                        logging.error("Country %s specified more than once",
                                      country)
                    else:
                        self._rules[country] = rule_collection
            elif state == ST_RULE:
                true_column = 1 + current_indent

                def cell_error_prefix():
                    return "Sheet \"%s\", %s: " % \
                        (ws.name, WorksheetBase.cell_name(row, true_column))

                m = re.match(
                    r"^\s*((?P<g2>2(\.4)?G(Hz)?)|(?P<g5>5G(Hz)?)|" +
                    r"((?P<l>\d+)(-(?P<h>\d+))?))\s*$",
                    ws.get_cell(row, true_column))
                fatal_if(
                    not m,
                    "%sInvalid band/range specification" % cell_error_prefix())
                if m.group("g2"):
                    band = ClmCompiler.Band._2
                    channel_range = None
                elif m.group("g5"):
                    band = ClmCompiler.Band._5
                    channel_range = None
                else:
                    channel_range = \
                        (int(m.group("l")),
                         int(m.group("h") if m.group("h")
                             else m.group("l")))
                    band = ClmCompiler.Band._2 if channel_range[0] <= 14 \
                        else ClmCompiler.Band._5
                rule_collection.setdefault(band, {}).setdefault(channel_range,
                                                                {})
                true_column += 1
                signature_text = ws.get_cell(row, true_column)
                column_signature = \
                    ColumnSignature(signature_text, cell_error_prefix())
                if column_signature in rule_collection[band][channel_range]:
                    logging.error("%sRule already defined", row_error_prefix)
                    continue

                rule_info = ColumnRules.RuleInfo(signature_text, row)
                rule_collection[band][channel_range][column_signature] = \
                    rule_info
                true_column += 1
                duplicating_rate_groups = set()
                for rate in ws.get_cell(row, true_column).split(","):
                    valid = False
                    m = re.search(
                        r"^\s*(?P<type>DSSS|OFDM|MCS|VHT)" +
                        r"((?P<idx_l>\d+)(-(?P<idx_h>\d+))?)?" +
                        r"((SS|x)(?P<ss>\d))?\s*" +
                        r"(\(\s*(?P<exp>CDD|STBC|TXBF|SDM|SPEXP)\s*\))?\s*$",
                        rate, re.IGNORECASE)
                    if (not m) or \
                            ((m.group("type").upper() in ("MCS", "VHT")) !=
                             bool(m.group("idx_l"))) or \
                            ((m.group("type").upper() == "VHT") !=
                             bool(m.group("ss"))):
                        logging.error("%sRate not recognized: \"%s\"",
                                      cell_error_prefix(), rate.strip())
                        continue
                    for s, v in [("DSSS", ClmCompiler.RateInfo.RateType.DSSS),
                                 ("OFDM", ClmCompiler.RateInfo.RateType.OFDM),
                                 ("MCS", ClmCompiler.RateInfo.RateType.MCS),
                                 ("VHT", ClmCompiler.RateInfo.RateType.VHT)]:
                        if m.group("type").upper() == s:
                            rate_type = v
                            break
                    if m.group("idx_l"):
                        rate_idx_l = int(m.group("idx_l"))
                        rate_idx_h = int(m.group("idx_h")) \
                            if m.group("idx_h") else rate_idx_l
                    if m.group("ss"):
                        vht_ss = int(m.group("ss"))
                    is_cdd = m.group("exp") == "CDD"
                    is_stbc = m.group("exp") == "STBC"
                    is_txbf = m.group("exp") == "TXBF"
                    for rate_info in \
                            ClmCompiler.RatesInfo.get_rates_by_chains(
                                column_signature.chains):
                        if (not rate_info.singular) or \
                                (rate_info.rate_type != rate_type):
                            continue
                        if rate_type in (ClmCompiler.RateInfo.RateType.MCS,
                                         ClmCompiler.RateInfo.RateType.VHT):
                            if (rate_info.modulation_index < rate_idx_l) or \
                                    (rate_info.modulation_index > rate_idx_h):
                                continue
                            if rate_type == ClmCompiler.RateInfo.RateType.VHT:
                                if vht_ss != \
                                        int(re.match(r"^VHT\d+SS(\d+)",
                                                     rate_info.name).group(1)):
                                    continue
                        skip = True
                        for flag, ss in [(is_cdd, "CDD"), (is_stbc, "STBC"),
                                         (is_txbf, "TXBF")]:
                            if flag != (ss in rate_info.name):
                                break
                        else:
                            skip = False
                        if skip:
                            continue
                        rule_info.add_rate(
                            rate_info,
                            duplicating_rate_groups=duplicating_rate_groups)
                        valid = True
                    if not valid:
                        logging.error(
                            "%sNo rates matching \"%s\" specification",
                            cell_error_prefix(), rate)
                if duplicating_rate_groups:
                    logging.error(
                        "%sFollowing rate groups defined more than once: %s",
                        cell_error_prefix(),
                        ", ".join(duplicating_rate_groups))

    def get_rates(self, sheet_name, band, channel, column_signature,
                  used_rules=None):
        """ Returns rates for given country cell

        Arguments:
        sheet_name       -- Country sheet name
        band             -- Band
        channel          -- Channel number
        column_signature -- Signature of column heading
        used_rules       -- Optional output parameter. None or dictionary where
                            information about retrieved rules is being put to.
                            Has the same structure as _rules, but without first
                            and last levels: indexed by band, then by channel
                            ranges (for whole-band rules this level contains
                            the only index: None), leaves are column signature
                            sets
        Returns Tuple (rates, rule_found). First element is rate set to use
        (taken either from rule or from column signature), second is boolean -
        True if rates were taken from rule, false if from column signature
        """
        sheet_rules = self._rules.get(sheet_name.strip())
        if not sheet_rules:
            return (column_signature.default_rates, False)
        band_rules = sheet_rules.get(band)
        if not band_rules:
            return (column_signature.default_rates, False)
        used_channel_range = None
        if None in band_rules:
            channel_rules = band_rules[None]
        else:
            for channel_range in band_rules:
                if channel_range[0] <= channel <= channel_range[1]:
                    channel_rules = band_rules[channel_range]
                    used_channel_range = channel_range
                    break
            else:
                return (column_signature.default_rates, False)
        rule_info = channel_rules.get(column_signature)
        if not rule_info:
            return (column_signature.default_rates, False)
        if used_rules is not None:
            used_rules.setdefault(band, {}).\
                setdefault(used_channel_range, set()).add(column_signature)
        return (rule_info.rates, True)

    def empty(self):
        """ True if no rules defined (probably Rules sheet is absent) """
        return len(self._rules) == 0

    def has_rules_for(self, sheet_name):
        """ True if there are rules for given country sheet name """
        return sheet_name.strip() in self._rules

    def get_unused_rules(self, sheet_name, used_rules):
        """ Returns list of data about rules, not used for given page

        Arguments:
        sheet_name -- Country sheet name
        used_rules -- Dictionary containing information of used rules -
                      accumulated in previous calls of get_rates()
        Returns List of (rule_heading, definition_row) tuples. First element is
        a text of rule column signature, second element is row on Rule page
        that contains the rule
        """
        sheet_rules = self._rules.get(sheet_name.strip())
        if not sheet_rules:
            return []
        ret = []
        for band, band_rules in sheet_rules.iteritems():
            for chan_range, chan_range_rules in band_rules.iteritems():
                for column_signature, rule_info in \
                        chan_range_rules.iteritems():
                    if (band in used_rules) and \
                            (chan_range in used_rules[band]) and \
                            (column_signature in used_rules[band][chan_range]):
                        continue
                    rule_info = sheet_rules[band][chan_range][column_signature]
                    ret.append((rule_info.signature_text, rule_info.row))
        return ret


class OlympicLimits(object):
    """ Reads power limits spreadsheet and verifies regions against it

    Private attributes:
    _workbook_filename                  -- Spreadsheet file name
    _sheet_chantype_channel_power_rates -- Maps sheet names to channel types to
                                           channel numbers to power limits to
                                           sets of rates. Disabled limits
                                           encoded as None
    """
    def __init__(self, workbook, required_sheet_names):
        """ Constructor

        Arguments:
        workbook             -- Workbook object
        required_sheet_names -- Sheet names to retrieve
        """
        self._workbook_filename = workbook.filename
        ST_COMMENT, ST_HEADING, ST_CHANNEL = range(3)
        self._sheet_chantype_channel_power_rates = OrderedDict()
        rule_pages = [rsn for rsn in ColumnRules.RULE_SHEET_NAMES
                      if rsn in workbook.sheet_names]
        if len(rule_pages) == 0:
            logging.warning(
                "Rules' page not found. Column headers will be used to " +
                "determine rate names")
        elif len(rule_pages) > 1:
            fatal("More than one rule page found: \"%s\"" %
                  "\", \"".join(rule_pages))
        column_rules = ColumnRules(
            workbook.sheet_by_name(rule_pages[0]) if rule_pages else None)
        for sheet_name in workbook.sheet_names:
            if sheet_name not in required_sheet_names:
                continue
            using_rules = column_rules.has_rules_for(sheet_name)
            if (not column_rules.empty()) and (not using_rules):
                logging.warning(
                    "Sheet \"%s\": No column rules defined found. Default " +
                    "ratesets will be assumed", sheet_name.strip())
            ws = workbook.sheet_by_name(sheet_name)
            state = ST_COMMENT
            column_signatures = {}
            bandwidth = None
            chan_type = None
            current_indent = 0
            used_rules = {}
            unruly_cells = OrderedDict()
            heading_row = 0
            for row in range(1, ws.nrows + 1):
                error_prefix = "Sheet \"%s\", row %d: " % (sheet_name, row)
                for new_indent in range(ws.ncols):
                    if ws.get_cell(row, 1 + new_indent):
                        break
                else:
                    new_indent = current_indent
                if state == ST_COMMENT:
                    if ws.get_cell(row, 1 + new_indent) == "Channel":
                        state = ST_HEADING
                        current_indent = new_indent
                    elif re.match(r"^\d+(\.\d*)?$",
                                  ws.get_cell(row, 1 + new_indent)):
                        fatal("%sChannel number not preceded by heading row" %
                              error_prefix)
                elif state == ST_HEADING:
                    if re.match(r"^\d+(\.\d*)?$",
                                ws.get_cell(row, 1 + current_indent)):
                        state = ST_CHANNEL
                    else:
                        fatal("%sHeading row not followed by channel row" %
                              error_prefix)
                elif state == ST_CHANNEL:
                    if re.match(r"^\d+(\.\d*)?$",
                                ws.get_cell(row, 1 + current_indent)):
                        state = ST_CHANNEL
                    elif ws.get_cell(row, 1 + new_indent) == "Channel":
                        state = ST_HEADING
                        current_indent = new_indent
                    else:
                        state = ST_COMMENT
                if state == ST_HEADING:
                    column_signatures = {}
                    bandwidth = None
                    chan_type = None
                    for col in range(3, ws.ncols + 1 - current_indent):
                        true_column = col + current_indent
                        cell = ws.get_cell(row, true_column)
                        if cell in ("", "Frequency"):
                            break
                        column_signature = \
                            ColumnSignature(
                                cell,
                                "Sheet \"%s\", %s: " %
                                (sheet_name,
                                 WorksheetBase.cell_name(row, true_column)))
                        if bandwidth is None:
                            bandwidth = column_signature.bw
                        elif bandwidth != column_signature.bw:
                            fatal(("%sBandwidth in column %s not same as " +
                                   "in column %s") %
                                  (error_prefix,
                                   WorksheetBase.col_num_to_alpha(true_column),
                                   WorksheetBase.col_num_to_alpha(
                                       3 + current_indent)))
                        fatal_if(not column_signature.default_rates,
                                 "%sNo matching rates for column %s" %
                                 (error_prefix,
                                  WorksheetBase.col_num_to_alpha(true_column)))
                        column_signatures[col] = column_signature
                        heading_row = row
                elif state == ST_CHANNEL:
                    channel = int(float(ws.get_cell(row, 1 + current_indent)))
                    band = ClmCompiler.Band._2 if channel <= 14 \
                        else ClmCompiler.Band._5
                    if chan_type is None:
                        chan_type = ClmCompiler.ChannelType.create(band,
                                                                   bandwidth)
                    elif band != chan_type.band:
                        fatal(("%sDifferent rows use channels from " +
                               "different bands") % error_prefix)
                    for col in range(3, max(column_signatures.keys()) + 1):
                        cell = ws.get_cell(row, col + current_indent).\
                            replace(",", ".")
                        if re.match(r"^(\+|-)?\d+(\.\d+)?$", cell):
                            power = float(cell)
                        elif re.match(
                                r"^(N/?A|Not Supported|Do not support|N/?S|" +
                                r"DISABLED)$",
                                cell, re.IGNORECASE):
                            power = None
                        else:
                            fatal(("%sUnrecognized power value \"%s\" in " +
                                   "column %s") %
                                  (error_prefix, cell,
                                   WorksheetBase.col_num_to_alpha(col)))
                        rates, rule_found = \
                            column_rules.get_rates(
                                sheet_name=sheet_name, band=band,
                                channel=channel,
                                column_signature=column_signatures[col],
                                used_rules=used_rules)
                        self._sheet_chantype_channel_power_rates.\
                            setdefault(sheet_name, OrderedDict()).\
                            setdefault(chan_type, OrderedDict()).\
                            setdefault(channel, OrderedDict()).\
                            setdefault(power, set()).update(rates)
                        if using_rules and (not rule_found):
                            unruly_cells.setdefault(
                                (heading_row, col + current_indent,
                                 frozenset([rate_info.group
                                            for rate_info in rates])), set()).\
                                add(row)
            for rule, row in \
                    sorted(
                        column_rules.get_unused_rules(sheet_name=sheet_name,
                                                      used_rules=used_rules),
                        key=operator.itemgetter(1)):
                logging.warning(
                    "Sheet \"%s\": Rule \"%s\" defined in row %d of rule " +
                    "page not in use",
                    sheet_name.strip(), rule, row)
            for (heading_row, col, rate_groups), rows in \
                    unruly_cells.iteritems():
                logging.warning(
                    "Sheet \"%s\": No rule for following cells: %s. " +
                    "Following default set of rates, that corresponds to " +
                    "heading defined in %s, is used: %s",
                    sheet_name.strip(),
                    ", ".join(WorksheetBase.cell_name(row=row, column=col)
                              for row in sorted(rows)),
                    WorksheetBase.cell_name(row=heading_row, column=col),
                    ", ".join(rate_groups))

    def check(self, clm_data, regrevs, sheets_to_countries):
        """ Checks limits

        Arguments:
        clm_data            -- ClmCompiler.ClmData object that contains CLM
                               data to check
        regrevs             -- List of regrevs
        sheets_to_countries -- Maps sheet names to country codes
        """
        def format_power(power_or_none):
            return "disabled" if power_or_none is None else \
                ("%gdBm" % power_or_none)
        # channels_valid = True
        # for sheet in self._sheet_chantype_channel_power_rates:
        #     for chan_type in self._sheet_chantype_channel_power_rates[sheet]:
        #         for channel in self._sheet_chantype_channel_power_rates[
        #                 sheet][chan_type]:
        #             if clm_data.valid_channels. \
        #                     get_valid_channel(chan_type, channel) is None:
        #                 channels_valid = False
        #                 logging.error(
        #                     "Sheet %s uses invalid %sM %sG channel %d" %
        #                     (sheet,
        #                      ClmCompiler.Bandwidth.name[chan_type.bandwidth],
        #                      ClmCompiler.Band.name[chan_type.band], channel))
        # if not channels_valid:
        #     sys.exit(1)
        for regrev in regrevs:
            for sheet in sheets_to_countries:
                fatal_if(sheet not in self._sheet_chantype_channel_power_rates,
                         ("\"%s\" does not contain properly filled sheet " +
                          "\"%s\"") % (self._workbook_filename, sheet))
                for cc in sheets_to_countries[sheet]:
                    ccrev = ClmCompiler.CcRev(cc, regrev)
                    region = clm_data.get_region(ccrev)
                    if region is None:
                        logging.error(
                            "Region %s not defined in CLM XML", ccrev)
                        continue
                    used_powers = {}
                    power_discrepancies = OrderedDict()
                    for chan_type in self._sheet_chantype_channel_power_rates[
                            sheet]:
                        for channel in \
                            self._sheet_chantype_channel_power_rates[
                                sheet][chan_type]:
                            rpd = clm_data.get_rate_power_dict(
                                region, channel, chan_type, extended=True)
                            for desired_power, rates in \
                                    self._sheet_chantype_channel_power_rates[
                                        sheet][chan_type][channel].iteritems():
                                violating_groups = OrderedDict()
                                for rate in rates:
                                    if not region.loc_reg_caps.check_rate(
                                            rate):
                                        continue
                                    if rate not in rpd:
                                        actual_power = None
                                    else:
                                        clm_power = \
                                            rpd[rate][1].values()[0]
                                        used_powers.setdefault(
                                            rpd[rate][0].loc_id,
                                            set()).add(clm_power)
                                        if clm_power.is_disabled:
                                            actual_power = None
                                        else:
                                            for actual_power in \
                                                    clm_power.powers_dbm:
                                                if actual_power == \
                                                        desired_power:
                                                    break
                                    if actual_power != desired_power:
                                        violating_groups.setdefault(
                                            rate.group, set()).add(
                                            actual_power)
                                if violating_groups:
                                    power_discrepancies.setdefault(
                                        (ccrev, chan_type.band,
                                         chan_type.bandwidth, desired_power,
                                         ", ".join(
                                             "%s(%s)" %
                                             (rg,
                                              ", ".join(format_power(p)
                                                        for p in
                                                        violating_groups[rg]))
                                             for rg in violating_groups)),
                                        set()).add(channel)
                    for (ccrev, band, bandwidth, power, rates), channels in \
                            power_discrepancies.iteritems():
                        logging.error(
                            "Spreadsheet power target for region %s on " +
                            "%sM %sG channel(s) %s is %s. It differs from " +
                            "XML power targets on the following rates: %s",
                            ccrev, ClmCompiler.Bandwidth.name[bandwidth],
                            ClmCompiler.Band.name[band],
                            ", ".join(str(c) for c in sorted(channels)),
                            format_power(power), rates)
                    unused_powers = {}
                    for loc_id in region.locale_ids.itervalues():
                        loc_data = clm_data.get_locale(loc_id)
                        if not loc_data:
                            continue
                        for power, rates in loc_data.chan_power.iteritems():
                            if power in used_powers.get(loc_id, set()):
                                continue
                            rate_groups = unused_powers.setdefault(power,
                                                                   set())
                            for rate in rates:
                                rate_groups.add(rate.group)
                    for power, rates in \
                            sorted(
                                unused_powers.iteritems(),
                                key=lambda kvp:
                                (kvp[0].channel_range.channel_type.band,
                                 kvp[0].channel_range.channel_type.
                                 bandwidth)):
                        ct = power.channel_range.channel_type
                        logging.warning(
                            "Region %s has power %s, on %sM %sG " +
                            "channel(s) %s (rates %s), not covered by " +
                            "spreadsheet",
                            ccrev, power.power_str(),
                            ClmCompiler.Bandwidth.name[ct.bandwidth],
                            ClmCompiler.Band.name[ct.band],
                            power.channel_range.range_str,
                            ", ".join(rates))
        if LoggingRegisteringFilter.num_errors or \
                LoggingRegisteringFilter.num_warnings:
            logging.info("%d errors, %d warnings found",
                         LoggingRegisteringFilter.num_errors,
                         LoggingRegisteringFilter.num_warnings)
            sys.exit(1)
        logging.info("No errors or warnings found")


def main(argv):
    """ Do the job:

    Arguments:
    argv -- Argument list without command name
    """
    argument_parser = MyArgumentParser(
        formatter_class=RawDescriptionHelpFormatter, fromfile_prefix_chars='@',
        description="Verifies regions in given CLM XML against Olympic " +
        "power limits spreadsheet",
        epilog="""Examples:
""")
    argument_parser.add_argument(
        "--regrev", metavar="REGREV", type=int, action="append", default=[],
        help="Regrevs for checked regions. This parameter is mandatory, " +
        "it may be specified several times")
    argument_parser.add_argument(
        "--sheet", metavar='"SHEET NAME;CC1,CC2,..."', action="append",
        default=[],
        help="Specifies sheet name and list of CCs that shall be verified " +
        "against this sheet. List may be comma separated or space " +
        " separated. This parameter is mandatory, it may be specified " +
        "several times. Don't forget to put  this parameter in double " +
        "quotes. Probably it is better be put to parameter file")
    argument_parser.add_argument(
        "xml", metavar="CLM_XML",
        help="CLM XML file that contains data being checked. This " +
        "parameter is mandatory")
    argument_parser.add_argument(
        "olymits", metavar="OLYMPIC_SPREADSHEET",
        help="Spreadsheet with Olympic limits. This parameter is mandatory")
    argument_parser.add_argument(
        "errors", metavar="ERRORS_FILE", nargs="?",
        help="File for error messages (optional).")

    if not argv:
        argument_parser.print_help()
        sys.exit(1)

    args = argument_parser.parse_args(argv)

    log_formatter = \
        logging.Formatter("%(filename)s: %(levelname)s: %(message)s")
    for handler in [logging.StreamHandler()] + \
            ([logging.FileHandler(args.errors, mode="w")]
             if args.errors else []):
        handler.setFormatter(log_formatter)
        logging.getLogger().addHandler(handler)
    logging.getLogger().addFilter(LoggingRegisteringFilter())
    logging.getLogger().setLevel(logging.INFO)
    fatal_if(not args.regrev, "No regrevs specified - nothing to check")

    workbook = open_workbook(args.olymits)
    sheet_to_ccs = OrderedDict()
    for arg in args.sheet:
        if ";" not in arg:
            fatal(
                "Sheet name and CC list shall be separated by ';' in \"%s\"" %
                arg)
        sheet, ccs = arg.split(";", 1)
        if sheet not in workbook.sheet_names:
            fatal("Sheet \"%s\" not found in spreadsheet" % sheet)
        for cc in re.split(r" |\t|,", ccs):
            if cc and (cc not in sheet_to_ccs.get(sheet, [])):
                sheet_to_ccs.setdefault(sheet, []).append(cc)

    olymits = OlympicLimits(workbook, sheet_to_ccs.keys())
    clm_data = ClmCompiler.ClmContainer(args.xml).fetch(
        ClmCompiler.FilterParams.filter_all(), False)
    olymits.check(clm_data, args.regrev, sheet_to_ccs)


# Standalone module starter
if __name__ == "__main__":
    main(sys.argv[1:])
